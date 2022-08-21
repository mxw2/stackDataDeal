# This is a sample Python script.
import time

import openpyxl
# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.

from openpyxl import Workbook


def write():
    book = Workbook()
    sheet = book.active

    sheet['A1'] = 56
    sheet['A2'] = 67

    new = time.strftime("%x")
    sheet['A3'] = new

    sheet.cell(row=3, column=4).value = 77

    # append
    rows = (
        (88, 46, 57),
        (89, 38, 12),
        (23, 59, 78),
        (56, 21, 98),
        (24, 18, 43),
        (34, 15, 67)
    )

    for row in rows:
        sheet.append(row)

    book.save("sample.xlsx")


def read():
    book = openpyxl.load_workbook('sample.xlsx')
    sheet = book.active
    a1 = sheet['A1']
    a2 = sheet["A2"]
    a3 = sheet.cell(row=3, column=4)

    print(a1.value)
    print(a2.value)
    print(a3.value)


def saveForRow():
    book = Workbook()
    sheet = book.active

    rows = (
        (88, 46, 57),
        (89, 38, 12),
        (23, 59, 78),
        (56, 21, 98),
        (24, 18, 43),
        (34, 15, 67)
    )

    for row in rows:
        sheet.append(row)
    # 列转行
    print("读取数据，将数据的列变成行")
    for row in sheet.iter_cols(min_row=1, min_col=1, max_row=6, max_col=3):
        for cell in row:
            print(cell.value, end=" ")
        print()

    print("读取数据，将数据的行变成列")
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
        for cell in row:
            print(cell.value, end="\n")
        print()

    book.save('iterbycols.xlsx')

    # sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3);

if __name__ == '__main__':
    saveForRow()
