from enum import Enum, unique
import openpyxl

@unique
class CalculateType(Enum):
    Year = 1
    OriginalData = 2  # 原始数据即可
    DivisionIncome = 3  # 除以营业收入
    # 经营活动产生利润 (income - business_cost - business_tax - selling_expenses - manage_expenses - develop_expenses)
    BusinessCreateProfit = 4
    GrossMargin = 5  # (1 - 成本占比) * 100 = 一个整数
    CommonTurnoverRate = 6  # 常用的周转率 B12/((B153+A153)/2) : B12 销售收入总额
    GoodsTurnoverRate = 7  # B19/((B112+A112)/2) : B19 成本总额


useful_years_max = 100000
useful_years_6 = 6
useful_years_4 = 4

ds_income_string = '营业收入'  # 12
ds_business_cost_string = '营业成本'  # 19
ds_business_tax_string = '营业税金及附加'  # 22
ds_selling_expenses_string = '销售费用'  # 23
ds_manage_expenses_string = '管理费用'  # 24
ds_develop_expenses_string = '研发费用'  # 25

# 提供给外界用的可变数组
models = []

# 使用步骤：
# 1.将要处理的excle放到python脚本同级目录下
# 2.修改python脚本中workbook_name为你的文件名称
# 3.确定ds_sheet表名是否相同
# 4.确定最后算出来的"共有x年"与你的一致，因为如果表格是你手动修改过，添加过超过原来最大列的cell值，算出来的就是有问题的
workbook_name = '黄山旅游08-18.xlsx'
# prepare data
book = openpyxl.load_workbook(workbook_name)
# 数据源sheet
ds_sheet = book["利润表,资产负债表,现金流量表3"]
# 数据源开始年的start char
ds_start_year_index_char = 'B'
# 数据源中 key:value
ds_first_column_dictionary = {}

# ************************** sheet 相关 **************************************


def create_result_sheet():
    # 判断是否存在，没有的话立刻创建，用于保存结果
    result_sheet_name = '结果'
    if result_sheet_name in book.sheetnames:
        # remove表格
        result_sheet = book[result_sheet_name]
        book.remove(result_sheet)
        print('remove掉已有的result sheet')
    else:
        print('没有[结果]sheet，需要手动新建')
    # 创建一张新的sheet, 用于存放处理完毕的数据
    return book.create_sheet(result_sheet_name, 1)


def create_first_column_dictionary():
    print("Maximum row: {0}".format(ds_sheet.max_row))
    # key:
    for row in range(ds_sheet.max_row):
        # print('current row ' + str(row + 1))
        current_row_string = str(row + 1)
        cell = ds_sheet["A" + current_row_string]
        if cell.value is None:
            continue
        elif len(cell.value) > 0:
            value = cell.value.strip()
            ds_first_column_dictionary[value] = current_row_string
            print('key:' + value + ', value:' + current_row_string)
        else:
            continue


def save_result_sheet():
    # 保存下
    book.save(workbook_name)


# ************************** ColumnModel 相关 **************************************


class ColumnModel:
    # useful_years 显示几年，默认10年
    def __init__(self, name, ds_row_index_string, ds_row_content, text_color, calculate_type, useful_years):
        self.name = name
        self.ds_row_index_string = ds_row_index_string
        self.ds_row_content = ds_row_content
        self.text_color = text_color
        self.calculate_type = calculate_type
        self.useful_years = useful_years


def ds_row_for_key(ds_row_content):
    assert len(ds_row_content) > 0, "字典必须大于0"
    if ds_first_column_dictionary.__contains__(ds_row_content):
        ds_row_index_string = ds_first_column_dictionary[ds_row_content]
        # print('find ds_content ' + ds_row_content + ', ds_row ' + ds_row_index_string)
        return ds_row_index_string
    else:
        assert False, ds_row_content + '没有找到row在ds_sheet'
        return '-1'


def create_column_model(name, ds_row_content, text_color, calculate_type, useful_years):
    assert len(ds_first_column_dictionary) > 0, "字典必须大于0"
    # default -1 有些就是不需要index
    ds_row_index_string = '-1'
    if ds_row_content is not None:
        ds_row_index_string = ds_row_for_key(ds_row_content)
        # print('find ds_content ' + ds_row_content + ', ds_row ' + ds_row_index_string)

    model = ColumnModel(name, ds_row_index_string, ds_row_content, text_color, calculate_type, useful_years)
    models.append(model)


def want_to_deal_with_data_source():
    # 因为有些具体科目在execl表格row位置不同，所以统一用文字搜索的方式获取，更加准确，cell中文字有空格，要先删除空格在匹配
    # ********************** 具体年数 **********************
    # 5
    create_column_model('年份', '报告参数', None, CalculateType.Year, useful_years_max)

    # ********************** 营业收入与经营活动现金流入图 **********************
    # 12
    create_column_model('营业收入（亿元）',
                        ds_income_string,
                        "00ff00",
                        CalculateType.OriginalData,
                        useful_years_max)
    # 244
    create_column_model('经营活动现金流入（亿元）',
                        '销售商品、提供劳务收到的现金',
                        "00ff00",
                        CalculateType.OriginalData,
                        useful_years_max)

    # ********************** 净利润现金净流对比分析 **********************
    # 62
    create_column_model('净利润',
                        '六、净利润',
                        "ffff00",
                        CalculateType.OriginalData,
                        useful_years_max)
    # 275
    create_column_model('经营活动产生的现金流量净额',
                        '经营活动产生的现金流量净额',
                        "ffff00",
                        CalculateType.OriginalData,
                        useful_years_max)

    # ********************** 历年现金流情况 **********************
    # 275
    create_column_model('经营活动净额(亿元)',
                        '经营活动产生的现金流量净额',
                        "ff0000",
                        CalculateType.OriginalData,
                        useful_years_6)
    # 318
    create_column_model('筹资净额（亿元）',
                        '筹资活动产生的现金流量净额',
                        "ff0000",
                        CalculateType.OriginalData,
                        useful_years_6)
    # 297
    create_column_model('投资净额（亿元）',
                        '投资活动产生的现金流量净额',
                        "ff0000",
                        CalculateType.OriginalData,
                        useful_years_6)

    # ********************** 历年收入成本构成(%) **********************
    # 19
    create_column_model('营业成本',
                        ds_business_cost_string,
                        "aaff00",
                        CalculateType.DivisionIncome,
                        useful_years_4)
    # 22
    create_column_model('营业税金及附加',
                        ds_business_tax_string,
                        "aaff00",
                        CalculateType.DivisionIncome,
                        useful_years_4)
    # 23
    create_column_model('销售费用',
                        ds_selling_expenses_string,
                        "aaff00",
                        CalculateType.DivisionIncome,
                        useful_years_4)
    # 24
    create_column_model('管理费用',
                        ds_manage_expenses_string,
                        "aaff00",
                        CalculateType.DivisionIncome,
                        useful_years_4)
    # 25
    create_column_model('研发费用',
                        ds_develop_expenses_string,
                        "aaff00",
                        CalculateType.DivisionIncome,
                        useful_years_4)
    # 通过计算的出来结果
    create_column_model('经营活动产生利润',
                        None,
                        "aaff00",
                        CalculateType.BusinessCreateProfit,
                        useful_years_4)

    # 历年毛利率(%) 通过计算得出来结果
    create_column_model('销售毛利率',
                        None,
                        "00ffaa",
                        CalculateType.GrossMargin,
                        useful_years_4)

    # 主要资产周转率（单位：次）
    # 153
    create_column_model('总资产周转率（次）',
                        '资产总计',
                        "00ccaa",
                        CalculateType.CommonTurnoverRate,
                        useful_years_4)
    # 95
    create_column_model('应收账款周转率（次）',
                        '应收票据及应收账款',
                        "00ccaa",
                        CalculateType.CommonTurnoverRate,
                        useful_years_4)
    # 112
    create_column_model('存货周转率（次）',
                        '存货',
                        "00ccaa",
                        CalculateType.GoodsTurnoverRate,
                        useful_years_4)
    # 133
    create_column_model('固定资产周转率（次）',
                        '固定资产',
                        "00ccaa",
                        CalculateType.CommonTurnoverRate,
                        useful_years_4)

    # ********************** 历年资产堆积图 **********************
    # 86
    create_column_model('货币资金（亿元）',
                        '货币资金',
                        "bbccaa",
                        CalculateType.OriginalData,
                        useful_years_max)
    # 112
    create_column_model('存货（亿元）',
                        '存货',
                        "bbccaa",
                        CalculateType.OriginalData,
                        useful_years_max)
    # 95
    create_column_model('应收票据及应收账款（亿元）',
                        '应收票据及应收账款',
                        "bbccaa",
                        CalculateType.OriginalData,
                        useful_years_max)
    # 133
    create_column_model('固定资产（亿元）',
                        '固定资产',
                        "bbccaa",
                        CalculateType.OriginalData,
                        useful_years_max)
    # 134
    create_column_model('在建工程（亿元）',
                        '在建工程',
                        "bbccaa",
                        CalculateType.OriginalData,
                        useful_years_max)
    # 128
    create_column_model('可供出售的金融资产（亿元）',
                        '可供出售金融资产',
                        "bbccaa",
                        CalculateType.OriginalData,
                        useful_years_max)

    # ********************** 历年负债堆积图 **********************
    # 164
    create_column_model('应付票据及应付账款（亿元）',
                        '应付票据及应付账款',
                        "bb2233",
                        CalculateType.OriginalData,
                        useful_years_max)
    # 172
    create_column_model('应交税费（亿元）',
                        '应交税费',
                        "bb2233",
                        CalculateType.OriginalData,
                        useful_years_max)
    # 173
    create_column_model('其他应付款（亿元）',
                        '其他应付款合计',
                        "bb2233",
                        CalculateType.OriginalData,
                        useful_years_max)

    return models

