from column_model import book, save_result_sheet
from openpyxl.chart import *
# from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import *

# ds_company_name =
ds_sheet = book["分业务收入占比总毛利占比"]
pie_company_name = ds_sheet['A1'].value
max_row = format(ds_sheet.max_row)
print("Maximum column: {0}".format(ds_sheet.max_row))


def create_all_business_income_chart(result_sheet, title, ds_column, target_position):
    chart = PieChart()
    chart.dLbls = DataLabelList()
    # chart.dLbls.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)], bodyPr=rot)
    chart.dLbls.showPercent = True

    # 营业收入
    values = Reference(result_sheet, min_col=ds_column, min_row=2, max_col=ds_column, max_row=max_row)
    # true 表示数据中不包含横向的titles()
    chart.add_data(data=values, titles_from_data=False)
    # 时间
    cats = Reference(result_sheet, min_col=1, min_row=2, max_col=1, max_row=max_row)
    chart.set_categories(cats)
    # 注释
    # s0 = chart.series[0]
    # s0.tx = SeriesLabel()
    # s0.tx.value = '净利润'
    # s0.graphicalProperties.solidFill = 'FF0000'
    # s1.dLbls.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)], bodyPr=rot)

    # s1 = chart.series[1]
    # s1.tx = SeriesLabel()
    # s1.tx.value = '经营活动现金净额'
    # s1.graphicalProperties.solidFill = '0938F7'
    chart.width = 16
    chart.height = 12
    chart.title = title
    result_sheet.add_chart(chart, target_position)


def cell_val(r, c):
    return ds_sheet.cell(row=r, column=c).value


# 创建营业收入占比饼图
create_all_business_income_chart(ds_sheet,
                                 '2019年' + pie_company_name + '各业务收入占比图',
                                 2,
                                 'A15')

gross_sum = 0
for m in range(2, int(max_row) + 1): # 2-6都要执行
    # 收入-成本
    gross = cell_val(m, 2) - cell_val(m, 3)
    gross_sum += gross

# 第二次遍历获取毛利占比总毛利百分比
for m in range(2, int(max_row) + 1):
    percent = ((cell_val(m, 2) - cell_val(m, 3)) / gross_sum) * 100
    # 扩大100倍
    ds_sheet.cell(row=m, column=4).value = percent
    print('m = ', m, '重写数据cell, percent = ', percent)

# 创建毛利占比百分比图
create_all_business_income_chart(ds_sheet,
                                 '2019年' + pie_company_name + '各业务毛利/总毛利占比图',
                                 4,
                                 'H15')
save_result_sheet()



