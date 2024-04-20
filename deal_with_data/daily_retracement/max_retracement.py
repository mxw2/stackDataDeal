print('开始统计当前股票的最大回撤')

# ********************************** 常量 **********************************

# 一周5个交易日
target_days = 5

# f1fe4-ssf8u 1998-2017
# 2o2bi-e3f13 2017-2024

from daily_price_info import DailyPriceInfo
import openpyxl
import math

workbook_name = 'f1fe4-ssf8u.xlsx'
# prepare data
book = openpyxl.load_workbook(workbook_name)
# 数据源sheet
ds_sheet = book["f1fe4-ssf8u"]

# 从第二行开始读数据
row_index_for_start_read = 2
column_index_for_date = 'A'
column_index_for_max_value = 'D'
column_index_for_min_value = 'E'

# 时间排序好的价格模型
price_infos = []

max_loss_percent: float = 0.0


# ********************************** 读取表内容 **********************************
def read_data():
    # 读取数据源表，并且做若干判断，保证数据位置都是正确的

    print("Maximum row: {0}, 有效行数量:{1}".format(ds_sheet.max_row, ds_sheet.max_row - 1))
    print("Maximum column: {0}".format(ds_sheet.max_column))

    # 遍历每一行数据
    # range(5):  # 从0到4
    # range(1, 6):  # 从1开始，到5结束（5会被包含在内）
    for i in range(2, ds_sheet.max_row + 1):
        date = ds_sheet[column_index_for_date + str(i)].value
        min_value = ds_sheet[column_index_for_min_value + str(i)].value
        max_value = ds_sheet[column_index_for_max_value + str(i)].value
        price_info = DailyPriceInfo(date, max_value, min_value)
        price_infos.append(price_info)

    # 按照时间由远到近的排序
    price_infos.reverse()


def max_retracement():
    # 统计最大亏损
    max_loss = 0
    global max_loss_percent
    # 这里可以使用2个info持有
    max_loss_start_price_info = None
    max_loss_end_price_info = None

    # 遍历每个价格 & 求出连续5天最大回撤
    for i in range(len(price_infos)):
        current_price_info = price_infos[i]
        # 【压力测试】买在当前最高价、触碰后4日最低价
        current_price = current_price_info.max_value
        for j in range(i + 1, i + target_days):
            # 特殊处理，保证j < price_infos.count
            if j >= len(price_infos):
                # print("遍历价格数组越界打印：i :{0}, j :{1}，continue".format(i, j))
                continue
            other_price_info = price_infos[j]
            # 使用最小值，便于压力测试，必须问问券商保证金随股价变化公式
            other_price = other_price_info.min_value
            if other_price < current_price:
                temp_loss = round(other_price - current_price, 2)
                temp_loss_percent = round(temp_loss / current_price, 4)
                # 这里不可以使用损失的钱数比较，要采用百分比比较才可以
                if temp_loss_percent < max_loss_percent:
                    max_loss = temp_loss
                    max_loss_percent = temp_loss_percent
                    max_loss_start_price_info = current_price_info
                    max_loss_end_price_info = other_price_info
                    # 给current_price_info缓存最大回撤百分比
                    current_price_info.retracement_percent = temp_loss_percent
                    current_price_info.debug_log(i)


    debug_log()

    # 打印最后的结果哈
    print("---------------------")
    describe_str = "压力测试:【最大值-最小值】\n"
    max_loss_str = f"最大损失: {str(max_loss)}\n"
    max_loss_percent_str = f"最大损失百分比: {str(round(max_loss_percent * 100, 2))}% \n"
    max_loss_start_str = f"开始日:{max_loss_start_price_info.date},最高价: {str(max_loss_start_price_info.max_value)}\n"
    max_loss_end_str = f"损失日: {max_loss_end_price_info.date},最低价: {str(max_loss_end_price_info.min_value)}\n"
    print(describe_str + max_loss_str + max_loss_percent_str + max_loss_start_str + max_loss_end_str)


def loss_distributions():
    global max_loss_percent
    # 如果最大亏损为-3.7%，则需要4个位置
    # 向上取整
    scale = 100
    count = math.ceil(-max_loss_percent * scale)
    loss_percent_distributes = [0] * count
    for i in range(len(price_infos)):
        current_price_info = price_infos[i]
        # 向下取整, 这里不太严谨，必须排除0的情况
        if current_price_info.retracement_percent == 0:
            continue
        index = math.floor(-current_price_info.retracement_percent * scale)
        print(f" 存储到数组的 index: {index}, count = {count}, 回撤 = {current_price_info.retracement_percent}")
        loss_percent_distributes[index] += 1

    print("---------------------")
    print(f"损失天数: {len(loss_percent_distributes)}")
    # 统计 & 打印
    for i in range(count):
        if loss_percent_distributes[i] > 0:
            print(
                f"损失范围: {-i}% 到 {-i - 1}%, 占比: {round(loss_percent_distributes[i] / len(price_infos) * 100, 2)}%, 数量: {loss_percent_distributes[i]}")


def debug_log():
    print("---------------------")
    for i in range(len(price_infos)):
        price_info = price_infos[i]
        # 打印调试
        # price_info.debug_log()


# ********************************** 启动 **********************************
if __name__ == '__main__':
    read_data()
    max_retracement()
    loss_distributions()
