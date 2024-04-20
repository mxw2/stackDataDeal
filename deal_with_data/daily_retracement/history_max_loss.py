print('开始统计当前股票的最大回撤')

# ********************************** 常量 **********************************

# 一周5个交易日
target_days = 5

# f1fe4-ssf8u 1998-2017
# 2o2bi-e3f13 2017-2024

from price_info import PriceInfo
import openpyxl
import math

workbook_name = 'AAPL_11_24.xlsx'
# prepare data
book = openpyxl.load_workbook(workbook_name)
# 数据源sheet
ds_sheet = book["AAPL_11_24"]

# 从第二行开始读数据
row_index_for_start_read = 2
column_index_for_date = 'A'
column_index_for_max_value = 'D'
column_index_for_min_value = 'E'

# 时间排序好的价格模型
price_infos = []

# 统计全局的最大损失值，和每一天for 4个寻找的max做对比，最大的放到这里即可
history_max_loss_price_info: PriceInfo = None


# ********************************** 读取表内容 **********************************
def read_data():
    # 读取数据源表，并且做若干判断，保证数据位置都是正确的
    print("Maximum row: {0}, 有效行数量:{1}".format(ds_sheet.max_row, ds_sheet.max_row - 1))
    print("Maximum column: {0}".format(ds_sheet.max_column))

    # 遍历每一行数据
    # range(5):  # 从0到4
    # range(1, 6):  # 从1开始，到5结束（5会被包含在内）
    for i in range(2, ds_sheet.max_row + 1):
        date = ds_sheet[column_index_for_date + str(i)].value
        min_value = ds_sheet[column_index_for_min_value + str(i)].value
        max_value = ds_sheet[column_index_for_max_value + str(i)].value
        price_info = PriceInfo(date, max_value, min_value)
        price_infos.append(price_info)

    # 按照时间由远到近的排序
    price_infos.reverse()


def max_loss():
    # 统计最大亏损
    global history_max_loss_price_info

    # 遍历每个价格 & 求出连续5天最大回撤
    for i in range(len(price_infos)):
        current_price_info = price_infos[i]
        # 【压力测试】买在当前最高价、触碰后4日最低价
        current_price = current_price_info.max_value

        for j in range(i + 1, i + target_days):
            # 特殊处理，保证j < price_infos.count
            if j >= len(price_infos):
                # print("遍历价格数组越界打印：i :{0}, j :{1}，continue".format(i, j))
                continue
            temp_price_info = price_infos[j]
            # 使用最小值，便于压力测试，必须问问券商保证金随股价变化公式
            temp_price = temp_price_info.min_value

            if temp_price < current_price:
                temp_loss_money = round(temp_price - current_price, 2)
                temp_loss_percent = round(temp_loss_money / current_price, 4)
                # 这里统计的是current_price_info对比4天寻找最大的损失并存储到current_price_info模型中
                # 最后整体再和历史上最大的损失百分比进行比较即可
                # 【注意】不可以使用损失的钱数比较，要采用百分比才可以
                if temp_loss_percent < current_price_info.loss_percent:
                    current_price_info.loss_percent = temp_loss_percent
                    current_price_info.loss_money = temp_loss_money
                    current_price_info.loss_price_info = temp_price_info

        # 此时current_price_info已经和后4天比较完毕最大损失
        # 和历史上最大损失进行比较
        if (history_max_loss_price_info is None or
                current_price_info.loss_percent < history_max_loss_price_info.loss_percent):
            # 【注意】此时回撤百分比是最小的，但是回撤价格不一定，因为股价不同，不能直接用价格比较
            history_max_loss_price_info = current_price_info

    debug_log()

    # 打印最后的结果哈
    print("---------------------")
    describe_str = "压力测试:【最大值-最小值】\n"
    max_loss_str = f"最大损失: {history_max_loss_price_info.loss_money}\n"
    max_loss_percent_str = f"最大损失百分比: {history_max_loss_price_info.loss_percent_str()}\n"
    max_loss_start_str = f"开始日:{history_max_loss_price_info.date}, 最高价: {history_max_loss_price_info.max_value}\n"
    history_max_loss_end_price_info = history_max_loss_price_info.loss_price_info
    max_loss_end_str = f"损失日: {history_max_loss_end_price_info.date},最低价: {history_max_loss_end_price_info.min_value}\n"
    print(describe_str + max_loss_str + max_loss_percent_str + max_loss_start_str + max_loss_end_str)


def loss_distributions():
    global history_max_loss_price_info

    assert history_max_loss_price_info is not None, "不可以为空哈"
    print("---------------------")
    # 向上取整
    count = math.ceil(-history_max_loss_price_info.loss_percent_expand_100())
    loss_percent_distributes = [0] * count
    loss_day_count = 0
    for i in range(len(price_infos)):
        current_price_info = price_infos[i]
        # 向下取整, 必须排除等于0的情况
        if current_price_info.loss_percent == 0:
            continue
        index = math.floor(-current_price_info.loss_percent_expand_100())
        print(f"【压测】在亏损分布数组的位置: {index}, 回撤百分比 = {current_price_info.loss_percent_str()}")
        loss_percent_distributes[index] += 1
        loss_day_count += 1

    print("---------------------")
    distribute_title_str = f"【压测】亏钱的范围 0% 到 -{len(loss_percent_distributes)}%, "
    sum_for_loss_day_percent_str = f"亏损天数占比: {round(loss_day_count / len(price_infos) * 100, 2)}%，"
    sum_for_trade_day_str = f"统计天数：{len(price_infos)},"
    sum_for_loss_day_str = f"亏损天数: {loss_day_count}"
    print(distribute_title_str + sum_for_loss_day_percent_str + sum_for_trade_day_str + sum_for_loss_day_str)

    # 统计 & 打印
    for i in range(count):
        num = loss_percent_distributes[i]
        if num > 0:
            loss_range_str = f"【压测】损失范围: {-i}% 到 {-i-1}%,"
            loss_range_percent_str = f" 占比: {round(num / len(price_infos) * 100, 2)}%,"
            loss_range_content_count_str = f"数量: {num}"
            print(loss_range_str + loss_range_percent_str + loss_range_content_count_str)


def debug_log():
    print("---------------------")
    for i in range(len(price_infos)):
        price_info = price_infos[i]
        # 打印调试
        price_info.debug_model(i)


# ********************************** 启动 **********************************
if __name__ == '__main__':
    read_data()
    max_loss()
    loss_distributions()