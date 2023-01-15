import os
import time
import openpyxl
from copy import copy
from setuptools.msvc import winreg


class ItemFilter:
    def __init__(self):
        # 数据源
        # True: 数组中n个词语各产生1张表, False: 数组中n个词语产生n张表
        self.only_one_result_workbook = False
        # 产品归属这一列中做过滤的可能数组
        self.ds_sheet_target_items = ['X86企业级']
        # 数据源sheet中具体列的名称,需要对他做过滤
        self.ds_sheet_target_column_name = '产品归属'
        self.ds_sheet_name = "信创表"
        self.workbook_name = 'ZKE30N-202301'
        self.workbook_name_xlsx = self.workbook_name + '.xlsx'
        self.book = None
        self.ds_sheet = None
        self.load_data_source_sheet()

        # 结果sheet
        self.result_work_book = None
        self.result_work_sheet = None

    def get_desktop_path(self):
        # 可能是windows获取桌面地址
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                             r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')
        return winreg.QueryValueEx(key, "Desktop")[0]
        '/Users/marsxinwang/Desktop/ZKE30N-202301.xlsx'
        # mac
        # return os.path.join(os.path.expanduser("~"), 'Desktop/')

    def get_desktop_filter_path(self):
        return self.get_desktop_path() + 'filter/'

    # ************************** 读取数据源 sheet **************************************
    def load_data_source_sheet(self):
        # prepare data
        path = self.get_desktop_path() + self.workbook_name_xlsx
        self.book = openpyxl.load_workbook(path, data_only=True)
        # 数据源sheet
        self.ds_sheet = self.book[self.ds_sheet_name]
        print("ds sheet max column: {0}".format(self.ds_sheet.max_column))
        print("ds sheet max row: {0}".format(self.ds_sheet.max_row))

    # ************************** 结果工作簿 **************************************

    # 创建新的过滤结果sheet & 将源数据第一行copy到过了结果sheet中
    def create_result_sheet(self):
        self.result_work_book = openpyxl.Workbook()
        self.result_work_sheet = self.result_work_book.active
        # 将数据源第一行数据copy到结果表格的第一行
        # openpyxls 选择整行后会变成一个cell组成的tuple对象，获取数据需要逐个获取：
        # https://stackoverflow.com/questions/23332259/copy-cell-style-openpyxl
        for cell in self.ds_sheet[1]:
            new_cell = self.result_work_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # 遍历设置的数组，放到同一个result_sheet
    def filter_all_target_item_to_one_result_sheet(self):
        # 从数据源第二行开始读取，看看哪个属于"ds_sheet_target_items"
        # "产品归属"这一列在G1
        # range（0， 5） 是[0, 1, 2, 3, 4]没有5
        for datasource_row in range(2, self.ds_sheet.max_row + 1):
            for target_item in self.ds_sheet_target_items:
                self.append_datasource_row_to_result_sheet_if_needed(datasource_row, target_item)

        print("结果 sheet 列数: {0}".format(self.result_work_sheet.max_column))
        print("结果 sheet 行数: {0}".format(self.result_work_sheet.max_row))

    # 给定一个词语，就去做判断即可
    def filter_a_item_to_one_result_sheet(self, target_item):
        for datasource_row in range(2, self.ds_sheet.max_row + 1):
            self.append_datasource_row_to_result_sheet_if_needed(datasource_row, target_item)

        print("结果 sheet 列数: {0}".format(self.result_work_sheet.max_column))
        print("结果 sheet 行数: {0}".format(self.result_work_sheet.max_row))

    def append_datasource_row_to_result_sheet_if_needed(self, datasource_row, target_item):
        # 内部是个公式，需要给转化成字符串"X86企业级"、"0"、"X86终端",load_workbook时候，已经做了只要数据的处理
        # 现在给强制转换成字符串
        ds_gx_str = str(self.ds_sheet['G' + str(datasource_row)].value)
        # 判断相等使用==，比较的是字符串内部，可以正常匹配；使用is是内存地址，如果2个字符串内存地址不一样，False
        if ds_gx_str == target_item:
            print('G' + str(datasource_row) + '值：' + ds_gx_str + ' is ' + target_item + '=' + str(ds_gx_str == target_item))
            # 将数据源这一行添加到filter_result_sheet中
            ds_current_row_data = [temp_item.value for temp_item in self.ds_sheet[datasource_row]]
            self.result_work_sheet.append(ds_current_row_data)

    def hidden_columns(self):
        self.result_work_sheet.column_dimensions['a'].hidden = 1
        self.result_work_sheet.column_dimensions['b'].hidden = 1
        self.result_work_sheet.column_dimensions['c'].hidden = 1

    def save_result_book(self, items_string):
        # 保存表格数据
        time_string = time.strftime("%Y-%m-%d %H:%M", time.localtime())
        book_name = time_string + '，过滤词:' + items_string + '，工作簿：' + self.workbook_name + '.xlsx'
        if not os.path.exists(self.get_desktop_filter_path()):
            os.makedirs(self.get_desktop_filter_path())
        path = self.get_desktop_filter_path() + book_name
        print('save path = ' + path)
        self.result_work_book.save(path)


if __name__ == '__main__':
    item_filter: ItemFilter = ItemFilter()
    if item_filter.only_one_result_workbook:
        # 只有1个表格
        item_filter.create_result_sheet()
        item_filter.filter_all_target_item_to_one_result_sheet()
        item_filter.hidden_columns()
        items_name = '_'.join(item_filter.ds_sheet_target_items)
        print('items_name = ' + items_name)
        item_filter.save_result_book(items_name)
    else:
        # 每个过滤词对应一个表格
        for item in item_filter.ds_sheet_target_items:
            item_filter.create_result_sheet()
            item_filter.filter_a_item_to_one_result_sheet(item)
            item_filter.hidden_columns()
            item_filter.save_result_book(item)