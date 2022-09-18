from openpyxl.styles import PatternFill
import column_model
from column_model import *
from openpyxl.chart import *
from openpyxl.chart.label import *
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, ColorChoice


def suitable_result_column(column):
    has_extra_a = column / 26
    # cell
    # remind_goods
    # days36_ = 4
    # print('has_extra_a = ' + str(has_extra_a))
    # 1/26= 0.0x
    if has_extra_a < 1:
        suitable_column_str = chr(column + ord('A'))
    else:
        deal_with_column = column % 26
        # 目前只处理A1 & AA1情况，BA1+不再考虑
        suitable_column_str = 'A' + chr(deal_with_column + ord('A'))
    # print('result.column = ' + str(column) + ',suitable.column = ' + suitable_column_str)
    return suitable_column_str


def read_data():
    # 读取数据源表，并且做若干判断，保证数据位置都是正确的
    print("Maximum column: {0}".format(ds_sheet.max_column))

    # 维护一个字典，保存"寻找的字符串":"row"
    create_first_column_dictionary()

    # 创建结果sheet
    result_sheet = create_result_sheet()
    print('result sheet获取成功 ', result_sheet)

    # result 会有（data_source_years + 1）列
    # 去除第一个A列
    data_source_years = ds_sheet.max_column - 1
    print('共有' + str(data_source_years) + '年')

    # 获取要处理的数据
    column_models = column_model.want_to_deal_with_data_source()

    # 1.从A1开始，先设置result的title row，不要和数据for-for混用，数值差1次
    print('先将title安置到第一行')
    # 超过26列，就有问题。[1了。。。
    for column in range(len(column_models)):
        model = column_models[column]
        result_cell_index = suitable_result_column(column) + '1'
        # print('title = ' + model.name + 'result_cell_index = ' + str(result_cell_index))
        result_sheet[result_cell_index] = model.name

    # 除以1亿
    yi_unit = 100000000.0

    # 开始填充result sheet，先遍历列，在遍历行
    # 2.从A2开始填充 column是0开始, A\B\C\D\E\F
    for column in range(len(column_models)):
        model = column_models[column]
        # row是从零开始的，但是在sheet是从第二行开始
        for row in range(data_source_years):
            # 有些cell不需要填充
            actually_userful_years = model.useful_years
            if model.useful_years == column_model.useful_years_max:
                actually_userful_years = data_source_years

            if data_source_years - row > actually_userful_years:
                continue
            # result每次从ds中获取数据的时候，result每次row + 1，datasource每次column + 1
            ds_cell_index = chr(row + ord(ds_start_year_index_char)) + model.ds_row_index_string
            # ds_cell_index = ds_row_for_key(ds_cell_string)
            result_cell_index = suitable_result_column(column) + str(row + 2)
            # print('双层for循环，转换index。ds_index[' + ds_cell_index + ':' + result_cell_index + ']result_index')
            # 逻辑判断 & 格式化数据
            if model.calculate_type == CalculateType.Year:
                year_str = ds_sheet[ds_cell_index]
                content = year_str.value.year
                # print('进入到CalculateType.Year:' + str(content))
            elif model.calculate_type == CalculateType.OriginalData:
                original_data = ds_sheet[ds_cell_index].value
                content = two_formate(original_data / yi_unit)
            elif model.calculate_type == CalculateType.DivisionIncome:
                original_data = ds_sheet[ds_cell_index].value
                # 获取当年的营业收入cell索引
                income_cell_index = chr(row + ord(ds_start_year_index_char)) + ds_row_for_key(ds_income_string)
                income_data = ds_sheet[income_cell_index].value
                # 先乘上100在除数，保证数据格式稍微好看点
                content = two_formate(original_data * 100 / income_data)
            elif model.calculate_type == CalculateType.BusinessCreateProfit:
                # 建议采用数组更好管理
                # 获取当年的营业收入cell索引
                income_cell_index = chr(row + ord(ds_start_year_index_char)) + \
                                    ds_row_for_key(ds_income_string)
                income_data = ds_sheet[income_cell_index].value
                remind_data = ds_sheet[income_cell_index].value

                add_items = [ds_business_cost_string,
                             ds_business_tax_string,
                             ds_selling_expenses_string,
                             ds_manage_expenses_string,
                             ds_develop_expenses_string]
                for item in add_items:
                    item_index = chr(row + ord(ds_start_year_index_char)) + ds_row_for_key(item)
                    remind_data -= ds_sheet[item_index].value

                # 先乘上100在除数，保证数据格式稍微好看点
                content = two_formate(remind_data * 100 / income_data)
            elif model.calculate_type == CalculateType.GrossMargin:
                # 获取当年的营业收入cell索引
                income_cell_index = chr(row + ord(ds_start_year_index_char)) + ds_row_for_key(ds_income_string)
                income_data = ds_sheet[income_cell_index].value
                remind_data = ds_sheet[income_cell_index].value
                business_cost_index = chr(row + ord(ds_start_year_index_char)) + ds_row_for_key(ds_business_cost_string)
                remind_data -= ds_sheet[business_cost_index].value
                # 先乘上100在除数，保证数据格式稍微好看点
                content = two_formate(remind_data * 100 / income_data)
            elif model.calculate_type == CalculateType.CommonTurnoverRate or model.calculate_type == CalculateType.GoodsTurnoverRate:
                # B12 / (B153 + A153) / 2
                # B19 / (B153 + A153) / 2
                if model.calculate_type == CalculateType.CommonTurnoverRate:
                    ds_target_string = ds_income_string
                else:
                    ds_target_string = ds_business_cost_string
                income_cell_index = chr(row + ord(ds_start_year_index_char)) + ds_row_for_key(ds_target_string)
                income_data = ds_sheet[income_cell_index].value
                # B153 & A153
                common_data = ds_sheet[ds_cell_index].value
                left_common_cell_index = chr(row + ord(ds_start_year_index_char) - 1) + model.ds_row_index_string
                left_common_data = ds_sheet[left_common_cell_index].value
                content = two_formate(income_data / ((left_common_data + common_data) / 2.0))
            elif model.calculate_type == CalculateType.InterestEarnLiabilities:
                interest_earn_liabilities = 0
                add_items = [ds_short_term_borrowing_string,
                             ds_long_term_borrowing_string,
                             ds_bonds_payable_string,
                             ds_non_current_liabilities_due_within_one_year_string,
                             ds_trading_financial_liabilities_string]
                for item in add_items:
                    item_index = chr(row + ord(ds_start_year_index_char)) + ds_row_for_key(item)
                    interest_earn_liabilities += ds_sheet[item_index].value
                content = two_formate(interest_earn_liabilities / yi_unit)
            else:
                content = ''

            result_sheet[result_cell_index] = content
            cell = result_sheet[result_cell_index]
            # 统一设置文字颜色等
            if model.text_color is not None:
                cell.fill = PatternFill("solid", fgColor=model.text_color)
    create_profit_and_business_net_cash_chart(result_sheet)
    create_income_and_business_cash_come_in_chart(result_sheet)
    create_cash_flow_chart(result_sheet)
    create_cost_structure_chart(result_sheet)
    create_gross_chart(result_sheet)
    create_turnover_rate_chart(result_sheet)
    create_asset_accumulation_chart(result_sheet)
    create_liabilities_accumulation_chart(result_sheet)
    save_result_sheet()
    # 建议搞一个营业收入 & 利润的表看看（科大讯飞是个雷）


def create_profit_and_business_net_cash_chart(result_sheet):
    chart = LineChart()
    chart.dLbls = DataLabelList()
    # chart.dLbls.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)], bodyPr=rot)
    chart.dLbls.showVal = True

    values = Reference(result_sheet, min_col=4, min_row=2, max_col=5, max_row=14)
    # true 表示数据中不包含横向的titles()
    chart.add_data(data=values, titles_from_data=False)
    # 时间
    cats = Reference(result_sheet, min_col=1, min_row=2, max_col=1, max_row=14)
    chart.set_categories(cats)
    # 注释
    s0 = chart.series[0]
    s0.tx = SeriesLabel()
    s0.tx.value = '净利润'
    s0.graphicalProperties.solidFill = 'FF0000'
    # s1.dLbls.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)], bodyPr=rot)

    s1 = chart.series[1]
    s1.tx = SeriesLabel()
    s1.tx.value = '经营活动现金净额'
    s1.graphicalProperties.solidFill = '0938F7'
    chart.width = 32
    chart.height = 16
    chart.title = ds_company_name + '历年净利润 & 经营活动现金净额（单位：亿元）'
    # chart.typ
    chart.x_axis.tickLblPos = 'low'
    result_sheet.add_chart(chart, 'AK27')
    # save_result_sheet()


def create_income_and_business_cash_come_in_chart(result_sheet):
    chart = LineChart()

    # set all labels
    # chart.dLbls = DataLabelList()
    # chart.dLbls.showVal = True

    # set axis label styles
    # chart1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)], bodyPr=rot)
    # chart1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)])

    values = Reference(result_sheet, min_col=2, min_row=2, max_col=3, max_row=14)
    # true 表示数据中不包含横向的titles()
    chart.add_data(data=values, titles_from_data=False)
    # 时间
    cats = Reference(result_sheet, min_col=1, min_row=2, max_col=1, max_row=14)
    chart.set_categories(cats)
    # 注释
    incom_charcter_properties = CharacterProperties(sz=1800, solidFill=ColorChoice(prstClr="blue"))
    incom_charcter_rot = openpyxl.drawing.text.RichTextProperties(vert='horz')
    s0 = chart.series[0]
    s0.marker.symbol = 'circle'
    s0.marker.size = 5
    s0.tx = SeriesLabel()
    s0.tx.value = '营业收入'
    s0.graphicalProperties.solidFill = 'FF0000'
    s0.dLbls = DataLabelList()
    s0.dLbls.dLblPos = 't'
    s0.dLbls.showVal = True
    s0.dLbls.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=incom_charcter_properties),
                                          endParaRPr=incom_charcter_properties)],
                             bodyPr=incom_charcter_rot)

    axis = CharacterProperties(sz=1800, solidFill=ColorChoice(prstClr="red"))
    rot = openpyxl.drawing.text.RichTextProperties(vert='horz')
    s1 = chart.series[1]
    s1.marker.symbol = 'circle'
    s1.marker.size = 5
    s1.tx = SeriesLabel()
    s1.tx.value = '经营活动现金流入'
    s1.graphicalProperties.solidFill = '0938F7'
    s1.dLbls = DataLabelList()
    s1.dLbls.showVal = True
    s1.dLbls.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)], bodyPr=rot)

    chart.width = 32
    chart.height = 16
    chart.title = ds_company_name + '历年营业收入 & 现金流入（单位：亿元）'
    # chart.typ
    chart.x_axis.tickLblPos = 'low'
    result_sheet.add_chart(chart, 'S27')
    # save_result_sheet()


def create_cash_flow_chart(result_sheet):
    # 创建图标
    chart = BarChart()
    # chart.title = Title(tx=Text(strRef=StrRef('ds_company_name')))
    chart.dLbls = DataLabelList()
    # chart.dLbls.showCatName = True  # 标签显示(显示第几列的 1,-23  显示不符合预期)
    chart.dLbls.showVal = True  # 数量显示

    # 1.先设置Y轴，看看每年的数据
    values = Reference(result_sheet, min_col=6, min_row=7, max_col=8, max_row=14)
    # s.marker.symbol = 'circle'
    # true 表示数据中不包含横向的titles()
    chart.add_data(data=values, titles_from_data=False)
    axis = CharacterProperties(sz=1200, solidFill=ColorChoice(prstClr="white"))
    rot = openpyxl.drawing.text.RichTextProperties(vert='horz')

    # 2.放到后边才能x轴正常出现年的时间
    cats = Reference(result_sheet, min_col=1, min_row=7, max_col=1, max_row=14)
    chart.set_categories(cats)
    # 修改表格背景的
    # chart.plot_area.graphicalProperties = GraphicalProperties(solidFill="999999")
    s0 = chart.series[0]
    # 设置线条颜色，不设置时默认为砖红色,其他颜色代码可参见 https://www.fontke.com/tool/rgb/00aaaa/
    # s2.graphicalProperties.line.solidFill = "FF5555"
    # s2.graphicalProperties.line.width = 30000  # 控制线条粗细
    # fill0 = PatternFillProperties()
    # fill0.background = ColorChoice(prstClr="red")
    s0.tx = SeriesLabel()
    s0.tx.value = '经营活动净额'
    s0.graphicalProperties.solidFill = 'FF0000'
    s0.dLbls = DataLabelList()
    s0.dLbls.showVal = True
    s0.dLbls.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)], bodyPr=rot)

    s1 = chart.series[1]
    s1.tx = SeriesLabel()
    s1.tx.value = '筹资活动净额'
    # fill1 = PatternFillProperties(prst='wave')
    # fill1.background = ColorChoice(prstClr="blue")
    s1.graphicalProperties.solidFill = '0938F7'
    s1.dLbls = DataLabelList()
    s1.dLbls.showVal = True
    s1.dLbls.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)], bodyPr=rot)

    s2 = chart.series[2]
    s2.tx = SeriesLabel()
    s2.tx.value = '投资活动净额'
    # fill2 = PatternFillProperties()
    # fill2.background = ColorChoice(prstClr="orange")
    s2.graphicalProperties.solidFill = 'EE9611'
    s2.dLbls = DataLabelList()
    s2.dLbls.showVal = True
    s2.dLbls.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)], bodyPr=rot)

    chart.width = 32
    chart.height = 16
    chart.type = 'col'
    # 自定义风格时候，不要用这个字段
    # chart.style = 15
    chart.shape = 4
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = ds_company_name + '历年现金流量净额（单位：亿元）'
    chart.x_axis.tickLblPos = 'low'
    result_sheet.add_chart(chart, 'A27')
    # save_result_sheet()


def create_cost_structure_chart(result_sheet):
    # 创建图标
    chart = BarChart()
    chart.dLbls = DataLabelList()
    # chart.dLbls.showCatName = True  # 标签显示(显示第几列的 1,-23  显示不符合预期)
    chart.dLbls.showVal = True  # 数量显示

    # 1.先设置Y轴，看看每年的数据
    values = Reference(result_sheet, min_col=9, min_row=9, max_col=14, max_row=14)
    # s.marker.symbol = 'circle'
    # true 表示数据中不包含横向的titles()
    chart.add_data(data=values, titles_from_data=False)

    # 2.放到后边才能x轴正常出现年的时间
    cats = Reference(result_sheet, min_col=1, min_row=9, max_col=1, max_row=14)
    chart.set_categories(cats)
    # 修改表格背景的
    # chart.plot_area.graphicalProperties = GraphicalProperties(solidFill="999999")
    s0 = chart.series[0]
    # 设置线条颜色，不设置时默认为砖红色,其他颜色代码可参见 https://www.fontke.com/tool/rgb/00aaaa/
    # s2.graphicalProperties.line.solidFill = "FF5555"
    # s2.graphicalProperties.line.width = 30000  # 控制线条粗细
    # fill0 = PatternFillProperties()
    # fill0.background = ColorChoice(prstClr="red")
    s0.tx = SeriesLabel()
    s0.tx.value = '营业成本'
    s0.graphicalProperties.solidFill = 'FF0000'

    s1 = chart.series[1]
    s1.tx = SeriesLabel()
    s1.tx.value = '营业税金及附加'
    # fill1 = PatternFillProperties(prst='wave')
    # fill1.background = ColorChoice(prstClr="blue")
    s1.graphicalProperties.solidFill = '0938F7'

    s2 = chart.series[2]
    s2.tx = SeriesLabel()
    s2.tx.value = '销售费用'
    # fill2 = PatternFillProperties()
    # fill2.background = ColorChoice(prstClr="orange")
    s2.graphicalProperties.solidFill = 'EE9611'

    s3 = chart.series[3]
    s3.tx = SeriesLabel()
    s3.tx.value = '管理费用'
    # s2.graphicalProperties.solidFill = 'EE9611'

    s4 = chart.series[4]
    s4.tx = SeriesLabel()
    s4.tx.value = '研发费用'
    # s2.graphicalProperties.solidFill = 'EE9611'

    s5 = chart.series[5]
    s5.tx = SeriesLabel()
    s5.tx.value = '经营活动产生利润'
    # s2.graphicalProperties.solidFill = 'EE9611'

    chart.width = 22
    chart.height = 16
    chart.type = 'col'
    # 自定义风格时候，不要用这个字段
    # chart.style = 15
    chart.shape = 4
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = ds_company_name + '历年收入成本构成（%）'
    chart.x_axis.tickLblPos = 'low'
    result_sheet.add_chart(chart, 'A59')


def create_gross_chart(result_sheet):
    chart = LineChart()
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True

    values = Reference(result_sheet, min_col=15, min_row=9, max_col=15, max_row=14)
    # true 表示数据中不包含横向的titles()
    chart.add_data(data=values, titles_from_data=False)
    # 时间
    cats = Reference(result_sheet, min_col=1, min_row=9, max_col=1, max_row=14)
    chart.set_categories(cats)
    # 注释
    s0 = chart.series[0]
    s0.tx = SeriesLabel()
    s0.tx.value = '毛利率'
    s0.graphicalProperties.solidFill = 'FF0000'

    # s1 = chart.series[1]
    # s1.tx = SeriesLabel()
    # s1.tx.value = '经营活动现金流入'
    # s1.graphicalProperties.solidFill = '0938F7'
    chart.width = 22
    chart.height = 16
    chart.title = ds_company_name + '历年毛利率（%）'
    # chart.typ
    chart.x_axis.tickLblPos = 'low'
    result_sheet.add_chart(chart, 'S59')
    # save_result_sheet()


def create_turnover_rate_chart(result_sheet):
    # 创建图标
    chart = LineChart()
    chart.dLbls = DataLabelList()
    # chart.dLbls.showCatName = True  # 标签显示(显示第几列的 1,-23  显示不符合预期)
    chart.dLbls.showVal = True  # 数量显示

    # 1.先设置Y轴，看看每年的数据
    values = Reference(result_sheet, min_col=16, min_row=9, max_col=19, max_row=14)
    # s.marker.symbol = 'circle'
    # true 表示数据中不包含横向的titles()
    chart.add_data(data=values, titles_from_data=False)

    # 2.放到后边才能x轴正常出现年的时间
    cats = Reference(result_sheet, min_col=1, min_row=9, max_col=1, max_row=14)
    chart.set_categories(cats)
    # 修改表格背景的
    # chart.plot_area.graphicalProperties = GraphicalProperties(solidFill="999999")
    s0 = chart.series[0]
    # 设置线条颜色，不设置时默认为砖红色,其他颜色代码可参见 https://www.fontke.com/tool/rgb/00aaaa/
    # s2.graphicalProperties.line.solidFill = "FF5555"
    # s2.graphicalProperties.line.width = 30000  # 控制线条粗细
    # fill0 = PatternFillProperties()
    # fill0.background = ColorChoice(prstClr="red")
    s0.tx = SeriesLabel()
    s0.tx.value = '总资产周转率（次）'
    s0.graphicalProperties.solidFill = 'FF0000'

    s1 = chart.series[1]
    s1.tx = SeriesLabel()
    s1.tx.value = '应收账款周转率（次）'
    # fill1 = PatternFillProperties(prst='wave')
    # fill1.background = ColorChoice(prstClr="blue")
    s1.graphicalProperties.solidFill = '0938F7'

    s2 = chart.series[2]
    s2.tx = SeriesLabel()
    s2.tx.value = '存货周转率（次）'
    # fill2 = PatternFillProperties()
    # fill2.background = ColorChoice(prstClr="orange")
    s2.graphicalProperties.solidFill = 'EE9611'

    s3 = chart.series[3]
    s3.tx = SeriesLabel()
    s3.tx.value = '固定资产周转率（次）'
    # s2.graphicalProperties.solidFill = 'EE9611'

    # s4 = chart.series[4]
    # s4.tx = SeriesLabel()
    # s4.tx.value = '研发费用'
    # # s2.graphicalProperties.solidFill = 'EE9611'
    #
    # s5 = chart.series[5]
    # s5.tx = SeriesLabel()
    # s5.tx.value = '经营活动产生利润'
    # # s2.graphicalProperties.solidFill = 'EE9611'

    chart.width = 22
    chart.height = 16
    chart.type = 'col'
    # 自定义风格时候，不要用这个字段
    # chart.style = 15
    chart.shape = 4
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = ds_company_name + '历年周转率（单位：次）'
    chart.x_axis.tickLblPos = 'low'
    result_sheet.add_chart(chart, 'AK59')


def create_asset_accumulation_chart(result_sheet):
    chart = AreaChart()
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True

    values = Reference(result_sheet, min_col=20, min_row=2, max_col=28, max_row=14)
    # true 表示数据中不包含横向的titles()
    chart.add_data(data=values, titles_from_data=False)
    # 时间
    cats = Reference(result_sheet, min_col=1, min_row=2, max_col=1, max_row=14)
    chart.set_categories(cats)
    # 注释
    s0 = chart.series[0]
    s0.tx = SeriesLabel()
    s0.tx.value = '货币资金'
    s0.graphicalProperties.solidFill = 'FF0000'

    s1 = chart.series[1]
    s1.tx = SeriesLabel()
    s1.tx.value = '存货'
    # s1.graphicalProperties.solidFill = '0938F7'

    s2 = chart.series[2]
    s2.tx = SeriesLabel()
    s2.tx.value = '应收票据及应收账款'
    # s2.graphicalProperties.solidFill = '0938F7'

    s3 = chart.series[3]
    s3.tx = SeriesLabel()
    s3.tx.value = '固定资产'
    # s3.graphicalProperties.solidFill = '0938F7'

    s4 = chart.series[4]
    s4.tx = SeriesLabel()
    s4.tx.value = '在建工程'
    # s4.graphicalProperties.solidFill = '0938F7'

    s5 = chart.series[5]
    s5.tx = SeriesLabel()
    s5.tx.value = '可供出售的金融资产'
    # s5.graphicalProperties.solidFill = '0938F7'

    s6 = chart.series[6]
    s6.tx = SeriesLabel()
    s6.tx.value = '商誉'
    # s6.graphicalProperties.solidFill = '0938F7'

    s7 = chart.series[7]
    s7.tx = SeriesLabel()
    s7.tx.value = '无形资产'
    # s7.graphicalProperties.solidFill = '0938F7' 好使

    s8 = chart.series[8]
    s8.tx = SeriesLabel()
    s8.tx.value = '其他流动资产'
    # s8.graphicalProperties.solidFill = '0938F7'

    chart.width = 32
    chart.height = 16
    chart.title = ds_company_name + '历年资产堆积图(单位：亿元)'
    chart.grouping = 'stacked'
    chart.x_axis.tickLblPos = 'low'
    result_sheet.add_chart(chart, 'A91')

    # save_result_sheet()


def create_liabilities_accumulation_chart(result_sheet):
    chart = AreaChart()
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True

    values = Reference(result_sheet, min_col=29, min_row=2, max_col=33, max_row=14)
    # true 表示数据中不包含横向的titles()
    chart.add_data(data=values, titles_from_data=False)
    # 时间
    cats = Reference(result_sheet, min_col=1, min_row=2, max_col=1, max_row=14)
    chart.set_categories(cats)
    # 注释
    s0 = chart.series[0]
    s0.tx = SeriesLabel()
    s0.tx.value = '应付票据及应付账款（亿元）'
    s0.graphicalProperties.solidFill = 'FF0000'

    s1 = chart.series[1]
    s1.tx = SeriesLabel()
    s1.tx.value = '应交税费（亿元）'
    # s1.graphicalProperties.solidFill = '0938F7'

    s2 = chart.series[2]
    s2.tx = SeriesLabel()
    s2.tx.value = '其他应付款（亿元）'
    # s2.graphicalProperties.solidFill = '0938F7'

    s3 = chart.series[3]
    s3.tx = SeriesLabel()
    s3.tx.value = '有息负债'
    # s3.graphicalProperties.solidFill = '0938F7'

    s4 = chart.series[4]
    s4.tx = SeriesLabel()
    s4.tx.value = '预收款项（亿元）'
    # s4.graphicalProperties.solidFill = '0938F7'

    # s5 = chart.series[5]
    # s5.tx = SeriesLabel()
    # s5.tx.value = '可供出售的金融资产'
    # # s5.graphicalProperties.solidFill = '0938F7'
    #
    # s6 = chart.series[6]
    # s6.tx = SeriesLabel()
    # s6.tx.value = '商誉'
    # # s6.graphicalProperties.solidFill = '0938F7'
    #
    # s7 = chart.series[7]
    # s7.tx = SeriesLabel()
    # s7.tx.value = '无形资产'
    # # s7.graphicalProperties.solidFill = '0938F7' 好使
    #
    # s8 = chart.series[8]
    # s8.tx = SeriesLabel()
    # s8.tx.value = '其他流动资产'
    # # s8.graphicalProperties.solidFill = '0938F7'

    chart.width = 32
    chart.height = 16
    chart.title = ds_company_name + '历年负债堆积图(单位：亿元)'
    chart.grouping = 'stacked'
    chart.x_axis.tickLblPos = 'low'
    result_sheet.add_chart(chart, 'S91')


# 返回两位小数数字
def two_formate(income):
    result_str = str(income).split('.')[0] + '.' + str(income).split('.')[1][:2]
    return float(result_str)


if __name__ == '__main__':
    read_data()
