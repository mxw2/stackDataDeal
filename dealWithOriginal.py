import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import column_model
from column_model import CalculateType

# preprare data
book = openpyxl.load_workbook('original.xlsx')
# 数据源sheet
ds_sheet = book["利润表,资产负债表,现金流量表1"]
# 数据源开始年的start char
start_year_index_char = 'B'
# 数据源开始年的end char
end_year_index_char = 'J'

def create_result_sheet():
    # 判断是否存在，没有的话立刻创建，用于保存结果
    result_sheet_name = '结果'
    if result_sheet_name in book.sheetnames:
        # remove表格
        result_sheet = book[result_sheet_name]
        book.remove(result_sheet)

    # 创建一张新的sheet, 用于存放处理完毕的数据
    return book.create_sheet(result_sheet_name, 1)

def read_data():
    # 查看下有数据单元格范围，表格不是特别准确，可能是我之前wb做了很多无用功导致的，以后在试试
    # print(data_source_sheet.dimensions)
    # print("Minimum row: {0}".format(data_source_sheet.min_row))
    # print("Maximum row: {0}".format(data_source_sheet.max_row))
    # print("Minimum column: {0}".format(data_source_sheet.min_column))
    # print("Maximum column: {0}".format(data_source_sheet.max_column))

    # 创建结果sheet
    result_sheet = create_result_sheet()

    # result 会有（data_source_years + 1）列
    data_source_years = ord(end_year_index_char) - ord(start_year_index_char) + 1
    print('共有' + str(data_source_years) + '年')

    # 获取开始年份
    # start_year_str = ds_sheet[start_year_index_char + '5']
    # print('start_year_str = ' + str(start_year_str.value.year))
    # result_sheet.append(first_row)

    # 获取要处理的数据
    column_models = column_model.want_to_deal_with_data_source()

    # 1.从A1开始，先设置result的title row，不要和数据for-for混用，数值差1次
    for column in range(len(column_models)):
        model = column_models[column]
        result_cell_index = chr(column + ord('A')) + '1'
        result_sheet[result_cell_index] = model.name

    # 除以1亿
    yi_unit = 100000000.0

    # 开始填充result sheet，先遍历列，在遍历行
    # 2.从A2开始填充 column是0开始, A\B\C\D\E\F
    for column in range(len(column_models)):
        model = column_models[column]
        # row是从零开始的，但是在sheet是从第二行开始
        for row in range(data_source_years):
            # result每次从ds中获取数据的时候，result每次row + 1，datasource每次column + 1
            ds_cell_index = chr(row + ord(start_year_index_char)) + str(model.ds_row_index)
            result_cell_index = chr(column + ord('A')) + str(row + 2)
            if model.calculate_type == CalculateType.Year:
                year_str = ds_sheet[ds_cell_index]
                result_sheet[result_cell_index] = year_str.value.year
            elif model.calculate_type == CalculateType.OriginalData:
                original_data = ds_sheet[ds_cell_index].value
                result_sheet[result_cell_index] = two_formate(original_data / yi_unit)
            elif model.calculate_type == CalculateType.DivisionIncome:
                original_data = ds_sheet[ds_cell_index].value
                # 获取当年的营业搜索cell索引
                income_cell_index = chr(row + ord(start_year_index_char)) + '12'
                income_data = ds_sheet[income_cell_index].value
                result_sheet[result_cell_index] = two_formate(original_data / income_data) * 100

    # 建议修改，数组都以1开始即可
    # 配置每一列数据
    # for index in range(data_source_years):
    #     # 数据源中开始遍历的列的index值
    #     ds_begin_colume_char = 'B'
    #     # 1列：拼写当前年
    #     result_sheet['A' + str(index + 2)] = str(start_year_str.value.year + index)
    #
    #     ds_current_column_char = ord(ds_begin_colume_char) + index
    #     # 2列：填写营业收入
    #     income_true = data_source_sheet[chr(ds_current_column_char) + ds_business_income_cell_row].value / yi_unit
    #     result_sheet['B' + str(index + 2)] = two_formate(income_true)
    #
    #     # 3列：营业收入
    #     income = data_source_sheet[chr(ds_current_column_char) + ds_cash_flow_cell_row].value / yi_unit
    #     result_sheet['C' + str(index + 2)] = two_formate(income)
    #
    #     # 4列:净利润
    #     income = data_source_sheet[chr(ds_current_column_char) + ds_net_profit_cell_row].value / yi_unit
    #     result_sheet['D' + str(index + 2)] = two_formate(income)
    #
    #     # 5列:经营活动产生的现金流量净额【要10年 & 要6年】
    #     income = data_source_sheet[chr(ds_current_column_char) + ds_net_cash_cell_row].value / yi_unit
    #     result_sheet['E' + str(index + 2)] = two_formate(income)
    #
    #     if data_source_years - index <= 6:
    #         cell = result_sheet['E' + str(index + 2)]
    #         cell.fill = PatternFill("solid", fgColor="ff0000")
    #         # 6列:筹资净额【只要6年的】
    #         income = data_source_sheet[chr(ds_current_column_char) + ds_net_money_in_cell_row].value / yi_unit
    #         result_sheet['F' + str(index + 2)] = two_formate(income)
    #         cell = result_sheet['F' + str(index + 2)]
    #         cell.fill = PatternFill("solid", fgColor="ff0000")
    #
    #         # 7列:投资净额【只要6年的】
    #         income = data_source_sheet[chr(ds_current_column_char) + ds_net_money_out_cell_row].value / yi_unit
    #         result_sheet['G' + str(index + 2)] = two_formate(income)
    #         cell = result_sheet['G' + str(index + 2)]
    #         cell.fill = PatternFill("solid", fgColor="ff0000")
    #
    #     if data_source_years - index <= 4:
    #         # 8列:'营业成本%'【只要4年】
    #         income = data_source_sheet[chr(ds_current_column_char) + ds_operating_cost_cell_row].value / yi_unit
    #         result_sheet['H' + str(index + 2)] = two_formate(income / income_true * 100)

    # 保存下
    book.save('original.xlsx')

# 返回两位小数数字
def two_formate(income):
    return str(income).split('.')[0] + '.' + str(income).split('.')[1][:2]
    # return int(result_str)

if __name__ == '__main__':
    read_data()